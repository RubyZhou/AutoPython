#----------------------------------------------------------
# Title   : writeExcel.py
# version : v1.0
# Create  : 2019/08/23 zhouyf
#
# Descripthion
#   使用 openpyxl 对 Excel 写操作
#
#----------------------------------------------------------

###########################################################
# list of functions.
###########################################################
#
# funciton name            function purpose
# -----------------------  -----------------------------
# CreateAndSaveWb()          创建和保存 Excel
#
###########################################################
# start of functions.
###########################################################

#
# Funtion : CreateAndSaveWb()
#    创建和保存 Excel
#
def CreateAndSaveWb():
    # 创建一个空的 workbook
    wb = openpyxl.Workbook()

    # 重命名 sheet名
    print(wb.get_sheet_names())
    sheet = wb.get_active_sheet()
    print(sheet.title)
    sheet.title = 'Spam Bacon Eggs Sheet'
    print(wb.get_sheet_names())

    # 保存工作簿
    wb.save('example_copy.xlsx')

#
# Funtion : CrtAndDelSheet()
#    创建和保存 Excel
#
def CrtAndDelSheet():
    wb = openpyxl.Workbook()
    print(wb.get_sheet_names())
    # 创建 sheet
    wb.create_sheet()
    wb.create_sheet(index=0, title='First Sheet')
    wb.create_sheet(index=2, title='Middle Sheet')
    print(wb.get_sheet_names())
    # 删除 sheet
    wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
    wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
    print(wb.get_sheet_names())



###########################################################
# end of functions.
###########################################################

###########################################################
#
#              MAIN BODY OF SCRIPT
#
###########################################################
import logging
import os
import openpyxl
logging.basicConfig(level=logging.DEBUG, format="[%(asctime)s][%(levelname)s] %(message)s")
logging.debug('Start of program')

CrtAndDelSheet()

# clean
if os.path.exists('example_copy.xlsx') == True:
    os.remove('example_copy.xlsx')