#----------------------------------------------------------
# Title   : useOpenpyxl.py
# version : v1.0
# Create  : 2019/10/21 zhouyf
#
# Descripthion
#   Excel 操作
#----------------------------------------------------------

###########################################################
# list of functions.
###########################################################
#
# funciton name            function purpose
# -----------------------  -----------------------------
# read_Workbook()          读取 workbook
# read_Worksheet()         读取 worksheet
# read_Cell()              读取 Cell
# get_maxRowAndCol()       获取表的大小
# get_Range()              遍历切片（sheet中某一范围）
#
###########################################################
#
# Funtion : read_Workbook()
#    读取 workbook
#
def read_Workbook() :
    wb = openpyxl.load_workbook('example.xlsx')
    print(type(wb))

#
# Funtion : read_Worksheet()
#    读取 worksheet
#
def read_Worksheet() :
    wb = openpyxl.load_workbook('example.xlsx')
    print(type(wb.get_sheet_names()))

    # 获取所有 sheet 名的 list
    for sheet in wb.get_sheet_names() :
        print(sheet)

    # 获取单个 sheet 对象
    sheet = wb.get_sheet_by_name('Sheet3')

    # 获取当前活动 sheet 对象
    anotherSheet = wb.get_active_sheet()

#
# Funtion : read_Cell()
#    读取 Cell
#
def read_Cell() :
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    # 获取单元格对象
    print(sheet['A1'])

    # 获取单元格值
    print(sheet['A1'].value)

    # 获取单元格的行和列信息
    c = sheet['B1']
    print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value) ## Row 1, Column B is Apples
    print('Cell ' + c.coordinate + ' is ' + c.value) ## Cell B1 is Apples
    print(sheet['C1'].value) ## 73

    # 按行和列信息获取 Cell对象
    c = sheet.cell(row=1, column=2)
    print(sheet.cell(row=1, column=2).value)
    print("---------------------------------")
    for i in range(1, 8, 2): # 遍历单元格
        print(i, sheet.cell(row=i, column=2).value)

#
# Funtion : get_maxRowAndCol()
#    获取表的大小
#
def get_maxRowAndCol():
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    print(sheet.max_row)
    print(sheet.max_column)

#
# Funtion : transferColumn()
#    列字母和数字之间的转换
#
def transferColumn() :
    from openpyxl.utils import get_column_letter, column_index_from_string

    # 列数字转字母
    print(get_column_letter(1))     # A
    print(get_column_letter(2))     # B
    print(get_column_letter(27))    # AA
    print(get_column_letter(900))   # AHP

    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')
    print(get_column_letter(sheet.max_column))

    # 列字母转数字
    print(column_index_from_string('A'))    # 1
    print(column_index_from_string('AA'))   # 27


#
# Funtion : get_Range()
#    遍历切片（sheet中某一范围）
#
def get_Range() :
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    # 放到一个元组里. (tuple() - 将字符串/list 转化成元组)
    print(tuple(sheet['A1':'C3']))  # (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>) 。。。)

    # 遍历切片(先遍历行对象，再遍历列)
    for rowOfCellObjects in sheet['A1':'C3']:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)
        print('--------- End of Row------------')

#
# Funtion : get_RowsAndColumns()
#    取整行和整列
#
def get_RowsAndColumns() :
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb.get_active_sheet()
    #print(tuple(sheet.columns)) # ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, ... (<Cell 'Sheet1'.B1>, ...)
    #print(tuple(sheet.rows))
    print(sheet["A"])
    print(sheet[1])

    for cellObj in sheet['A'] :
        print("行号:", cellObj.coordinate, " 值:", cellObj.value)


###########################################################
#
#              MAIN BODY OF SCRIPT
#
###########################################################
import openpyxl

get_RowsAndColumns()

