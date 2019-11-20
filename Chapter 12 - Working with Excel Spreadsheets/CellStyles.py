#----------------------------------------------------------
# Title   : CellStyles.py
# version : v1.0
# Create  : 2019/10/30 zhouyf
#
# Descripthion
#   字体和风格
#
#----------------------------------------------------------

###########################################################
# list of functions.
###########################################################
#
# funciton name            function purpose
# -----------------------  -----------------------------
# SetFont()                调整字体和风格
#
###########################################################
# start of functions.
###########################################################

#
# Funtion : SetFont()
#    调整字体和风格
#
def SetFont() :
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name('Sheet')
    italic24Font = Font(size=24, italic=True)
    sheet['A1'].font = italic24Font
    sheet['A1'].value = "Hello World!"
    wb.save('style.xlsx')

#
# Funtion : Set_RowCol_Size()
#    设置行高和列宽
#
def Set_RowCol_Size() :
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet['A1'] = 'Tall row'
    sheet['B2'] = 'Wide column'
    sheet.row_dimensions[1].height = 70
    sheet.column_dimensions['B'].width = 20
    wb.save('dimensions.xlsx')

#
# Funtion : MergeCells()
#    合并单元格
#
def MergeCells()  :
    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    sheet.merge_cells('A1:D3')
    sheet['A1'] = 'Twelve cells merged together.'

    sheet.merge_cells('C5:D5')
    sheet['C5'] = 'Two merged cells.'
    wb.save('merged.xlsx')


#
# Funtion : Chart()
#    使用图表
#
def Chart() :
    # 先清理
    clean()

    wb = openpyxl.Workbook()
    sheet = wb.get_active_sheet()
    for i in range(1, 11):
        sheet['A' + str(i)] = i

    # 矩形区域选择的单元格, 创建 Reference对象
    refObj = openpyxl.chart.Reference(sheet, min_row = 1, min_col = 1, max_row = 12, max_col = 1)

    # 通过传入 Reference对象, 创建 Series对象
    seriesObj = openpyxl.chart.Series(refObj, title='First series')

    # 将 Series对象 添加到 Chart对象
    chartObj = openpyxl.chart.BarChart()
    chartObj.append(seriesObj)

    # 设置： 图表标题
    chartObj.title = 'My Chart'

    # 将图表添加至 sheet 页从 C5 开始画
    sheet.add_chart(chartObj, 'C5')

    wb.save('sampleChart.xlsx')



#
# Funtion : clean()
#    清理
#
def clean() :
    import os
    files=["style.xlsx", "dimensions.xlsx",  "merged.xlsx", "sampleChart.xlsx"]
    i=0
    for fileName in  files:
        i=i+1
        if os.path.exists(fileName) == True :
            os.remove(fileName)
            print('[%-15s]已删除, [%d/%d]' %(fileName, i, len(files)))
        else:
            print('[%-15s]不存在, [%d/%d]' % (fileName, i, len(files)))

    # print("清理完成.")

###########################################################
# end of functions.
###########################################################



###########################################################
#
#              MAIN BODY OF SCRIPT
#
###########################################################
import openpyxl
from openpyxl.styles import Font


Chart()

















