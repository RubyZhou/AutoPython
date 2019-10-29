#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.


import openpyxl, os

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon' : 1.27
}


# Loop through the rows and update the prices.
updateNum=0
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
        updateNum = updateNum + 1

wb.save('updateProduceSales.xlsx')
print("一共更新了[%d]条数据", updateNum)

if True == os.path.exists('updateProduceSales.xlsx'):
    os.remove('updateProduceSales.xlsx')